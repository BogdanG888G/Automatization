from __future__ import annotations

import csv
import logging
import os
import re
import shutil
import subprocess
import tempfile
import time
from contextlib import closing
import pickle
from typing import Any, Dict, Iterable, Iterator, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd

import pyodbc
from sqlalchemy import create_engine, event, inspect, text
from sqlalchemy.engine import Engine

logger = logging.getLogger(__name__)

# =====================================================================================
# Connection defaults (override in DAG / runtime)
# =====================================================================================
DEFAULT_CONN_TEST = (
    "mssql+pyodbc://airflow_agent:123@host.docker.internal/Test"
    "?driver=ODBC+Driver+17+for+SQL+Server&Encrypt=yes&TrustServerCertificate=yes"
)


# =====================================================================================
# fast_executemany safeguard for SQLAlchemy Engine
# =====================================================================================

def enable_sqlserver_fast_executemany(engine: Engine) -> None:
    """Enable pyodbc fast_executemany *once* per Engine."""
    if getattr(engine, "_magnit_fast_exec_enabled", False):
        return

    @event.listens_for(engine, "before_cursor_execute")
    def _fast_exec(conn, cursor, statement, parameters, context, executemany):  # pragma: no cover - SQL hook
        if executemany and hasattr(cursor, "fast_executemany"):
            cursor.fast_executemany = True

    engine._magnit_fast_exec_enabled = True  # type: ignore[attr-defined]
    logger.debug("fast_executemany enabled for engine %s", engine)


# =====================================================================================
# Month dictionaries (ru)
# =====================================================================================
RU_MONTHS_FULL = [
    "январь", "февраль", "март", "апрель", "май", "июнь",
    "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь",
]
RU_MONTHS3 = [m[:3] for m in RU_MONTHS_FULL]
MAGNIT_MONTHS: Dict[str, int] = {m: i + 1 for i, m in enumerate(RU_MONTHS_FULL)}
MAGNIT_MONTHS.update({m: i + 1 for i, m in enumerate(RU_MONTHS3)})


# =====================================================================================
# Helpers
# =====================================================================================

def extract_magnit_metadata(source: str) -> Tuple[Optional[int], Optional[int]]:
    """Extract (month, year) from filename: `magnit_Март_2024.xlsx` → (3, 2024)."""
    tokens = re.findall(r"[a-zа-я]+|\d{4}", str(source).lower())
    year = next((int(t) for t in tokens if t.isdigit() and len(t) == 4), None)
    month = None
    for t in tokens:
        if t in MAGNIT_MONTHS:
            month = MAGNIT_MONTHS[t]
            break
    return month, year


def _stringify_header_cell(v: Any, i: int) -> str:
    if v is None:
        return f"none_{i}"
    s = str(v).strip()
    return s if s else f"none_{i}"


def normalize_magnit_columns(cols: Iterable[Any]) -> List[str]:
    """Normalize, dedupe, and RU→EN map column names."""
    normed = [re.sub(r"\s+", "_", str(c).strip().lower()) for c in cols]
    seen: Dict[str, int] = {}
    out: List[str] = []
    for c in normed:
        if c in seen:
            seen[c] += 1
            out.append(f"{c}_{seen[c]}")
        else:
            seen[c] = 0
            out.append(c)
    mapping = {
        "код_магазина": "store_code",
        "название_магазина": "store_name",
        "код_товара": "product_code",
        "наименование_товара": "product_name",
        "наименование": "product_name",
        "наименование_тп": "product_name",
        "количество": "quantity",
        "сумма": "amount",
        "дата": "date",
        "месяц": "month",
        "год": "year",
        "неделя": "week",
        "код_позиции": "position_code",
        "штриховой_код": "barcode",
        "продажи_в_шт.": "quantity_sold",
        "себестоимость_в_руб.": "cost_price",
        "продажи_в_руб.": "sales_amount",
    }
    out = [mapping.get(c, c) for c in out]
    return out


# =====================================================================================
# DDL inference (only used when raw_text_mode=False)
# =====================================================================================
NUMERIC_HINTS = {"quantity", "amount", "sales_amount", "cost_price", "quantity_sold"}
DATE_HINTS = {"date"}
INT_LIKE_SUFFIXES = {"_id", "_code", "_num", "_qty"}


def infer_sqlalchemy_type(col: str) -> str:
    c = col.lower()
    if c in NUMERIC_HINTS:
        return f"[{col}] DECIMAL(18, 2) NULL"
    if c in DATE_HINTS:
        return f"[{col}] DATE NULL"
    if any(c.endswith(sfx) for sfx in INT_LIKE_SUFFIXES):
        return f"[{col}] NVARCHAR(64) NULL"
    return f"[{col}] NVARCHAR(255) NULL"


def nvarcharmax(col: str) -> str:  # helper for raw_text_mode
    return f"[{col}] NVARCHAR(MAX) NULL"


# =====================================================================================
# File Readers
# =====================================================================================
class _FileReader:
    """Stream‑capable multi‑format reader (CSV / XLSX / XLSB)."""

    def __init__(
        self,
        sample_rows: int | None = 10_000,
        csv_sep: Optional[str] = None,
        csv_bad_lines: str = "error",
        xlsb_sheet_pattern: Optional[re.Pattern[str]] = None,
        xlsb_header_search_rows: int = 20,
        xlsx_read_only: bool = False,
    ) -> None:
        self.sample_rows = sample_rows
        self.csv_sep = csv_sep
        self.csv_bad_lines = csv_bad_lines
        self.xlsb_sheet_pattern = xlsb_sheet_pattern
        self.xlsb_header_search_rows = xlsb_header_search_rows
        self.xlsx_read_only = xlsx_read_only

    # ---------- CSV helpers ----------
    def _sniff_csv_sep(self, path: str) -> str:
        if self.csv_sep:
            return self.csv_sep
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            sample = f.read(128 * 1024)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            return dialect.delimiter
        except Exception:  # pragma: no cover
            counts = {d: sample.count(d) for d in [",", ";", "\t", "|"]}
            return max(counts, key=counts.get)

    def _csv_on_bad_lines(self) -> str:
        return self.csv_bad_lines if self.csv_bad_lines in {"error", "warn", "skip"} else "error"

    def sample_csv(self, path: str) -> pd.DataFrame:
        sep = self._sniff_csv_sep(path)
        try:
            return pd.read_csv(path, dtype="string", sep=sep, encoding="utf-8-sig", nrows=self.sample_rows,
                               low_memory=False, on_bad_lines=self._csv_on_bad_lines())
        except Exception as e:  # pragma: no cover
            logger.warning("CSV sample fallback (python engine): %s", e)
            return pd.read_csv(path, dtype=str, sep=sep, encoding="utf-8-sig", nrows=self.sample_rows,
                               low_memory=False, engine="python", on_bad_lines=self._csv_on_bad_lines())

    def iter_csv(self, path: str, chunksize: int) -> Iterator[pd.DataFrame]:
        sep = self._sniff_csv_sep(path)
        for chunk in pd.read_csv(path, dtype="string", sep=sep, encoding="utf-8-sig", chunksize=chunksize,
                                 low_memory=False, on_bad_lines=self._csv_on_bad_lines()):
            yield chunk

    # ---------- XLSX ----------
    def sample_xlsx(self, path: str) -> pd.DataFrame:
        return pd.read_excel(path, sheet_name=0, dtype="string", engine="openpyxl", nrows=self.sample_rows)

    def iter_xlsx(self, path: str, chunksize: int) -> Iterator[pd.DataFrame]:
        # read sheet->slice; if xlsx_read_only True we stream via openpyxl load_workbook(data_only=True, read_only=True)
        if not self.xlsx_read_only:
            sheets = pd.read_excel(path, sheet_name=None, dtype="string", engine="openpyxl")
            for sheet_name, df in sheets.items():
                if df.empty:
                    continue
                logger.debug("XLSX sheet %s rows=%s", sheet_name, len(df))
                for start in range(0, len(df), chunksize):
                    yield df.iloc[start:start + chunksize].copy()
        else:  # memory light path (slower per row but no giant DF)
            import openpyxl  # lazy
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_iter = ws.iter_rows(values_only=True)
                header = next(rows_iter, None)
                if not header:
                    continue
                header = [_stringify_header_cell(v, i) for i, v in enumerate(header)]
                batch: List[List[Any]] = []
                for row in rows_iter:
                    batch.append(list(row))
                    if len(batch) >= chunksize:
                        yield pd.DataFrame(batch, columns=header)
                        batch = []
                if batch:
                    yield pd.DataFrame(batch, columns=header)

    # ---------- XLSB ----------
    def _xlsb_detect_header(self, rows: Sequence[Sequence[Any]]) -> Tuple[int, List[str]]:
        for idx, row in enumerate(rows):
            non_null = [c for c in row if c not in (None, "", " ")]
            if len(non_null) >= 2:
                return idx, [_stringify_header_cell(v, i) for i, v in enumerate(row)]
        return 0, [_stringify_header_cell(v, i) for i, v in enumerate(rows[0])]

    def sample_xlsb(self, path: str) -> pd.DataFrame:
        import pyxlsb  # lazy
        with pyxlsb.open_workbook(path) as wb:  # pragma: no cover - IO heavy
            sheet_name = None
            for sn in wb.sheets:
                if not getattr(self, "xlsb_sheet_pattern", None) or self.xlsb_sheet_pattern.search(sn):
                    sheet_name = sn
                    break
            if sheet_name is None:
                raise ValueError("No XLSB sheets matched pattern.")
            with wb.get_sheet(sheet_name) as sheet:
                rows = []
                for i, row in enumerate(sheet.rows()):
                    rows.append([c.v for c in row])
                    if self.sample_rows is not None and i >= self.sample_rows:
                        break
        if not rows:
            return pd.DataFrame()
        hdr_idx, headers = self._xlsb_detect_header(rows[: self.xlsb_header_search_rows])
        data_rows = rows[hdr_idx + 1 :]
        return pd.DataFrame(data_rows, columns=headers)

    def iter_xlsb(self, path: str, chunksize: int) -> Iterator[pd.DataFrame]:
        import pyxlsb  # lazy
        with pyxlsb.open_workbook(path) as wb:  # pragma: no cover - IO heavy
            for sheet_name in wb.sheets:
                if getattr(self, "xlsb_sheet_pattern", None) and not self.xlsb_sheet_pattern.search(sheet_name):
                    logger.debug("Skipping XLSB sheet %s", sheet_name)
                    continue
                with wb.get_sheet(sheet_name) as sheet:
                    logger.debug("XLSB reading sheet %s", sheet_name)
                    row_iter = sheet.rows()
                    header_buf: List[List[Any]] = []
                    for _ in range(self.xlsb_header_search_rows):
                        try:
                            header_buf.append([c.v for c in next(row_iter)])
                        except StopIteration:
                            break
                    if not header_buf:
                        continue
                    hdr_idx, headers = self._xlsb_detect_header(header_buf)
                    data_buf = header_buf[hdr_idx + 1 :]
                    batch: List[List[Any]] = []
                    if data_buf:
                        batch.extend(data_buf)
                    for row in row_iter:
                        vals = [c.v for c in row]
                        if all(v in (None, "", " ") for v in vals):
                            continue
                        batch.append(vals)
                        if len(batch) >= chunksize:
                            yield pd.DataFrame(batch, columns=headers)
                            batch = []
                    if batch:
                        yield pd.DataFrame(batch, columns=headers)

    # ---------- unified ----------
    def sample(self, path: str) -> pd.DataFrame:
        pl = path.lower()
        if pl.endswith('.csv'):
            return self.sample_csv(path)
        if pl.endswith('.xlsx') or pl.endswith('.xls'):
            return self.sample_xlsx(path)
        if pl.endswith('.xlsb'):
            return self.sample_xlsb(path)
        raise ValueError(f"Unsupported format: {path}")

    def iter_chunks(self, path: str, chunksize: int) -> Iterator[pd.DataFrame]:
        pl = path.lower()
        if pl.endswith('.csv'):
            yield from self.iter_csv(path, chunksize)
        elif pl.endswith('.xlsx') or pl.endswith('.xls'):
            yield from self.iter_xlsx(path, chunksize)
        elif pl.endswith('.xlsb'):
            yield from self.iter_xlsb(path, chunksize)
        else:
            raise ValueError(f"Unsupported format: {path}")


# =====================================================================================
# MagnitTableProcessor (Turbo)
# =====================================================================================
class MagnitTableProcessor:
    """High‑volume loader for Magnit retail files.

    * Auto XLSX→CSV conversion (optional).
    * XLSB streaming support.
    * Bulk insert via pyodbc fast_executemany.
    * Raw text load mode (skip Python type coercion) for max speed.
    """

    DEFAULT_CONN_TEST = DEFAULT_CONN_TEST  # re‑export for convenience

    def __init__(
        self,
        engine: Engine,
        raw_schema: str = "raw",
        sample_rows: Optional[int] = 10_000,
        chunksize: int = 100_000,
        insert_batch_size: int = 50_000,
        create_table_if_missing: bool = True,
        dtype_overrides: Optional[Dict[str, str]] = None,
        add_sale_period_from_filename: bool = True,
        autocommit_batches: int = 5,
        raw_text_mode: bool = False,
        pyodbc_conn_str: Optional[str] = None,  # if provided, bypass engine.raw_connection
        convert_excel_to_csv: bool = True,      # for .xlsx only; xlsb streamed directly
        drop_temp_csv: bool = True,
        **file_reader_kwargs: Any,
    ) -> None:
        self.engine = engine
        self.raw_schema = raw_schema
        self.sample_rows = sample_rows
        self.chunksize = chunksize
        self.insert_batch_size = insert_batch_size
        self.create_table_if_missing = create_table_if_missing
        self.dtype_overrides = {k.lower(): v for k, v in (dtype_overrides or {}).items()}
        self.add_sale_period_from_filename = add_sale_period_from_filename
        self.autocommit_batches = autocommit_batches
        self.raw_text_mode = raw_text_mode
        self.pyodbc_conn_str = pyodbc_conn_str
        self.convert_excel_to_csv = convert_excel_to_csv
        self.drop_temp_csv = drop_temp_csv
        enable_sqlserver_fast_executemany(engine)
        self.reader = _FileReader(sample_rows=sample_rows, **file_reader_kwargs)
        self._load_enrichment_models()

    

    # ------------------------------------------------------------------
    # Core API
    # ------------------------------------------------------------------
    def process_file(self, path: str) -> str:
        start = time.time()
        fname = os.path.basename(path)
        table_name = self._table_name_from_filename(fname)
        fq_table = f"[{self.raw_schema}].[{table_name}]"
        logger.info("Processing file %s -> %s", fname, fq_table)

        # sale period once
        m: Optional[int] = None
        y: Optional[int] = None
        if self.add_sale_period_from_filename:
            m, y = extract_magnit_metadata(fname)

        # Possibly convert Excel→CSV for turbo path
        work_path = self._maybe_convert_excel_to_csv(path)
        is_temp = work_path != path
        try:
            # читаем первый чанк для создания таблицы
            reader = self.reader.iter_chunks(work_path, self.chunksize)
            first_chunk = next(reader)
            if first_chunk.empty:
                raise ValueError("No data in first chunk; aborting.")
            first_chunk.columns = normalize_magnit_columns(first_chunk.columns)
            first_chunk = self._enrich_product_data(first_chunk)
            if self.add_sale_period_from_filename and m and y:
                first_chunk = first_chunk.assign(sale_year=str(y), sale_month=f"{m:02d}")
            if not self.raw_text_mode:
                first_chunk = self._coerce_chunk_types(first_chunk)

            # создаём таблицу по enriched колонкам
            if self.create_table_if_missing:
                self._ensure_table(first_chunk, table_name)

            # открываем соединение на вставку
            total_rows = 0
            insert_conn, insert_cursor = self._get_insert_cursor()
            try:
                batch_rows = 0  # executemany batches counter for autocommit

                # вставляем первый чанк
                inserted, batch_rows = self._insert_chunk_fast(insert_cursor, first_chunk, fq_table, batch_rows)
                total_rows += inserted
                logger.info("Inserted %s rows (chunk 1, cumulative %s)", inserted, total_rows)

                # остальные чанки
                for i, chunk in enumerate(reader, start=2):
                    if chunk.empty:
                        continue
                    chunk.columns = normalize_magnit_columns(chunk.columns)
                    chunk = self._enrich_product_data(chunk)
                    if self.add_sale_period_from_filename and m and y and (
                        'sale_year' not in chunk.columns or 'sale_month' not in chunk.columns
                    ):
                        chunk = chunk.assign(sale_year=str(y), sale_month=f"{m:02d}")

                    if not self.raw_text_mode:
                        chunk = self._coerce_chunk_types(chunk)

                    inserted, batch_rows = self._insert_chunk_fast(insert_cursor, chunk, fq_table, batch_rows)
                    total_rows += inserted
                    logger.info("Inserted %s rows (chunk %s, cumulative %s)", inserted, i, total_rows)

                # final commit
                insert_conn.commit()
            finally:
                insert_cursor.close()
                insert_conn.close()

            dur = time.time() - start
            rps = total_rows / dur if dur > 0 else float("nan")
            logger.info("File %s loaded: %s rows in %.2fs (%.0f rows/s)", fname, total_rows, dur, rps)
            return table_name
        finally:
            if is_temp and self.drop_temp_csv:
                try:
                    os.remove(work_path)
                except OSError:
                    logger.warning("Could not remove temp CSV %s", work_path)

    def _load_enrichment_models(self):
        def load_model_and_vectorizer(model_name: str):
            with open(f"ml_models/product_enrichment/{model_name}_model.pkl", "rb") as f_model:
                model = pickle.load(f_model)
            with open(f"ml_models/product_enrichment/{model_name}_vectorizer.pkl", "rb") as f_vec:
                vectorizer = pickle.load(f_vec)
            return model, vectorizer

        self.brand_model, self.brand_vectorizer = load_model_and_vectorizer("brand")
        self.flavor_model, self.flavor_vectorizer = load_model_and_vectorizer("flavor")
        self.weight_model, self.weight_vectorizer = load_model_and_vectorizer("weight")
        self.type_model, self.type_vectorizer = load_model_and_vectorizer("type")

        # Адресные модели
        address_dir = "ml_models/product_enrichment"
        self.city_model, self.city_vectorizer = load_model_and_vectorizer("city", address_dir)
        self.region_model, self.region_vectorizer = load_model_and_vectorizer("region", address_dir)
        self.district_model, self.district_vectorizer = load_model_and_vectorizer("district", address_dir)
        self.branch_model, self.branch_vectorizer = load_model_and_vectorizer("branch", address_dir)

    
    def _enrich_product_data(self, df: pd.DataFrame) -> pd.DataFrame:
        if 'product_name' in df.columns:
            product_names = df['product_name'].fillna("")

            def predict(model, vectorizer):
                X_vec = vectorizer.transform(product_names)
                return model.predict(X_vec)

            df['brand_predicted'] = predict(self.brand_model, self.brand_vectorizer)
            df['flavor_predicted'] = predict(self.flavor_model, self.flavor_vectorizer)
            df['weight_predicted'] = predict(self.weight_model, self.weight_vectorizer)
            df['type_predicted'] = predict(self.type_model, self.type_vectorizer)

        # Адресное обогащение
        address_col_candidates = [c for c in df.columns if 'адрес' in c.lower() or 'ад' in c.lower() or 'ad' in c.lower()]
        if address_col_candidates:
            address_col = address_col_candidates[0]
            addresses = df[address_col].fillna("")

            def predict_address(model, vectorizer):
                X_vec = vectorizer.transform(addresses)
                return model.predict(X_vec)

            df['city_predicted'] = predict_address(self.city_model, self.city_vectorizer)
            df['region_predicted'] = predict_address(self.region_model, self.region_vectorizer)
            df['district_predicted'] = predict_address(self.district_model, self.district_vectorizer)
            df['branch_predicted'] = predict_address(self.branch_model, self.branch_vectorizer)

        return df

    # ------------------------------------------------------------------
    # Internals
    # ------------------------------------------------------------------
    @staticmethod
    def _table_name_from_filename(fname: str) -> str:
        base = os.path.splitext(fname)[0]
        return re.sub(r"\W+", "_", base.lower()).strip("_")

    def _ensure_table(self, df: pd.DataFrame, table_name: str) -> None:
        inspector = inspect(self.engine)
        if inspector.has_table(table_name, schema=self.raw_schema):
            logger.debug("Table %s.%s exists; skipping create", self.raw_schema, table_name)
            return

        cols: List[str] = []
        for col in df.columns:
            if self.raw_text_mode:
                dd = nvarcharmax(col)
            else:
                dd = self.dtype_overrides.get(col.lower()) or infer_sqlalchemy_type(col)
            cols.append(dd)
        create_sql = f"CREATE TABLE [{self.raw_schema}].[{table_name}] (" + ", ".join(cols) + ")"
        with self.engine.begin() as conn:
            conn.execute(text(create_sql))
        logger.info("Created table %s.%s", self.raw_schema, table_name)

    def _coerce_chunk_types(self, chunk: pd.DataFrame) -> pd.DataFrame:
        # Vectorized numeric/date coercion; skip if raw_text_mode
        for col in chunk.columns:
            lower = col.lower()
            if lower in NUMERIC_HINTS:
                s = chunk[col].astype("string")
                s = s.str.replace(" ", "", regex=False)
                s = s.str.replace(",", ".", regex=False)
                chunk[col] = pd.to_numeric(s, errors="coerce")
            elif lower in DATE_HINTS:
                chunk[col] = pd.to_datetime(chunk[col], errors="coerce").dt.date
        return chunk

    # ----- connection for insert -----
    def _get_insert_cursor(self) -> Tuple[Any, Any]:  # returns (conn, cursor)
        if self.pyodbc_conn_str:
            conn = pyodbc.connect(self.pyodbc_conn_str)
            conn.autocommit = False
            cursor = conn.cursor()
            cursor.fast_executemany = True
            return conn, cursor
        # else use engine raw_connection
        raw_conn = self.engine.raw_connection()  # DBAPI connection
        cursor = raw_conn.cursor()
        if hasattr(cursor, "fast_executemany"):
            cursor.fast_executemany = True
        return raw_conn, cursor

    # ----- fast insert chunk -----
    def _insert_chunk_fast(self, cursor: Any, chunk: pd.DataFrame, fq_table: str, batch_counter: int) -> Tuple[int, int]:
        cols = list(chunk.columns)
        # pyodbc expects Python sequences; values.tolist() ok; to_records() can produce numpy scalars not always safe
        data = chunk.where(pd.notna(chunk), None).values.tolist()  # convert NaN->None
        placeholders = ",".join(["?"] * len(cols))
        col_sql = ",".join(f"[{c}]" for c in cols)
        insert_sql = f"INSERT INTO {fq_table} ({col_sql}) VALUES ({placeholders})"

        inserted = 0
        # slice into insert_batch_size groups
        for start in range(0, len(data), self.insert_batch_size):
            part = data[start : start + self.insert_batch_size]
            cursor.executemany(insert_sql, part)
            inserted += len(part)
            batch_counter += 1
            if self.autocommit_batches and batch_counter >= self.autocommit_batches:
                # commit on underlying connection
                cursor.connection.commit()
                batch_counter = 0
        return inserted, batch_counter

    # ----- Excel→CSV convert (turbo path) -----
    def _maybe_convert_excel_to_csv(self, path: str) -> str:
        pl = path.lower()
        if not self.convert_excel_to_csv:
            return path
        if pl.endswith('.xlsx') or pl.endswith('.xls'):
            return self._convert_xlsx_cli_or_pandas(path)
        # do NOT auto‑convert xlsb; we'll stream via pyxlsb
        return path

    def _convert_xlsx_cli_or_pandas(self, xlsx_path: str) -> str:
        """Convert Excel to CSV; prefer external `xlsx2csv` tool for speed."""
        tmp_csv = tempfile.NamedTemporaryFile(delete=False, suffix=".csv").name
        # try xlsx2csv CLI
        try:
            subprocess.run(["xlsx2csv", "-d", ",", xlsx_path, tmp_csv], check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            logger.info("xlsx2csv converted %s -> %s", xlsx_path, tmp_csv)
            return tmp_csv
        except Exception as e:  # fallback
            logger.warning("xlsx2csv failed (%s); falling back to pandas.read_excel", e)
            # read all sheets, concat; WARNING memory heavy for giant >2M; but better than fail
            sheets = pd.read_excel(xlsx_path, sheet_name=None, dtype="string", engine="openpyxl")
            df = pd.concat(sheets.values(), ignore_index=True) if sheets else pd.DataFrame()
            df.to_csv(tmp_csv, index=False)
            logger.info("pandas converted %s -> %s", xlsx_path, tmp_csv)
            return tmp_csv


# =====================================================================================
# Backward‑compatible wrapper
# =====================================================================================

def create_magnit_table_and_upload(file_path: str, engine: Engine) -> str:
    proc = MagnitTableProcessor(engine=engine)
    return proc.process_file(file_path)


# =====================================================================================
# CLI smoke test
# =====================================================================================
if __name__ == "__main__":  # pragma: no cover - manual testing
    import argparse

    parser = argparse.ArgumentParser(description="Load Magnit file into SQL Server (turbo)")
    parser.add_argument("--conn", default=DEFAULT_CONN_TEST, help="SQLAlchemy URL")
    parser.add_argument("path", help="Input file path")
    parser.add_argument("--schema", default="raw")
    parser.add_argument("--chunksize", type=int, default=100_000)
    parser.add_argument("--batch", type=int, default=50_000)
    parser.add_argument("--raw-text", action="store_true", help="Load all cols as NVARCHAR(MAX)")
    parser.add_argument("--pyodbc", default=None, help="Optional raw pyodbc conn string (overrides engine)")
    args = parser.parse_args()

    eng = create_engine(args.conn, fast_executemany=True)
    proc = MagnitTableProcessor(
        eng,
        raw_schema=args.schema,
        chunksize=args.chunksize,
        insert_batch_size=args.batch,
        raw_text_mode=args.raw_text,
        pyodbc_conn_str=args.pyodbc,
    )
    tbl = proc.process_file(args.path)
    print(f"Loaded into {args.schema}.{tbl}")
