"""
create_table_and_upload_ashan.py  —  Ашан RAW loader (stream + chunk insert)

Основные цели:
- Поддержка ОЧЕНЬ больших файлов (до ~миллиона+ строк на лист).
- Потоковая обработка Excel/CSV: читаем чанками, не держим всё в памяти.
- Нормализация заголовков, маппинг колонок под канонические тех.имена.
- Извлечение периода (месяц/год) из имени файла ИЛИ из данных.
- Добавление отдельных колонок sale_year, sale_month и sale_date (dd-mm-yyyy; day=1 если нет в данных).
- Создание таблицы RAW (NVARCHAR(255) по умолчанию; override через TYPE_OVERRIDES).
- Авто-добавление новых колонок при встрече в следующих чанках.
- Быстрая вставка в MS SQL через pyodbc fast_executemany батчами.
- Подробное логирование по этапам и прогрессу.

Зависимости: pandas, numpy, sqlalchemy (engine), pyodbc, openpyxl.

Схема использования:
    from sqlalchemy import create_engine
    engine = create_engine('mssql+pyodbc://...')
    from create_table_and_upload_ashan import create_ashan_table_and_upload
    create_ashan_table_and_upload('/path/to/Ашан_сентябрь_2024.xlsx', engine)

Примечания:
- Все необработанные значения приводятся к строкам и обрезаются до 255 символов (RAW зона).
- Если в данных присутствует явная колонка даты (например, "дата продажи", "date", "dt", "period"), 
  будет предпринята попытка распарсить её и сформировать sale_date точнее, чем просто 1-е число месяца.
- Если день недоступен, в sale_date пишем первый день месяца ("YYYY-MM-01").
- sale_date хранится как NVARCHAR(10) в RAW (формат YYYY-MM-DD). В STAGE можно привести к DATE.

Настройки внизу файла см. класс AshanTableProcessor.
"""

from __future__ import annotations

import os
import re
import logging
import time
from contextlib import closing
from io import StringIO
from typing import Dict, Iterable, Iterator, List, Optional, Sequence, Tuple
import pickle

import numpy as np
import pandas as pd
import pyodbc
from openpyxl import load_workbook
from sqlalchemy import text, event
from sqlalchemy.engine import Engine

logger = logging.getLogger(__name__)

# ------------------------------------------------------------------ #
# Month dictionaries (рус / англ / сокращения / числовые шаблоны)   #
# ------------------------------------------------------------------ #
MONTHS_RU_FULL = {
    "январь": 1, "февраль": 2, "март": 3, "апрель": 4, "май": 5, "июнь": 6,
    "июль": 7, "август": 8, "сентябрь": 9, "октябрь": 10, "ноябрь": 11, "декабрь": 12,
}
MONTHS_RU_SHORT = {
    "янв": 1, "фев": 2, "мар": 3, "апр": 4, "май_": 5, "май": 5, "июн": 6,
    "июл": 7, "авг": 8, "сен": 9, "сент": 9, "окт": 10, "ноя": 11, "дек": 12,
}
MONTHS_EN_FULL = {
    "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12,
}
MONTHS_EN_SHORT = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "jun": 6, "jul": 7, "aug": 8,
    "sep": 9, "sept": 9, "oct": 10, "nov": 11, "dec": 12,
}

ASHAN_MONTHS: Dict[str, int] = {}
ASHAN_MONTHS.update(MONTHS_RU_FULL)
ASHAN_MONTHS.update(MONTHS_RU_SHORT)
ASHAN_MONTHS.update(MONTHS_EN_FULL)
ASHAN_MONTHS.update(MONTHS_EN_SHORT)

# ------------------------------------------------------------------ #
# Column canonical mapping (Ashan raw рус → нормализованное тех имя) #
# !!! Дополняй по мере поступления новых файлов.                     #
# ------------------------------------------------------------------ #
ASHAN_COLUMN_MAPPING: Dict[str, str] = {
    # Организационные / торговые единицы
    "магазин": "store",
    "код_магазина": "store_code",
    "город": "city",
    "регион": "region",
    "адрес": "address",
    # Товарные атрибуты
    "код_товара": "material",
    "код": "material",  # fallback
    "материал": "material",  # встречается в унифицированных выгрузках
    "наименование": "product_name",
    "бренд": "brand",
    "категория": "category",
    "подкатегория": "subcategory",
    # Метрики продаж
    "количество": "quantity",
    "qty": "quantity",
    "объем": "quantity",
    "оборот": "gross_turnover",
    "оборот_с_ндс": "gross_turnover",
    "себестоимость": "gross_cost",
    "средняя_цена": "avg_sell_price",
    "средняя_цена_продажи": "avg_sell_price",
    "средняя_цена_по_себестоимости": "avg_cost_price",
    # Периодность / даты
    "дата": "date_raw",  # исходная дата в данных
    "period": "date_raw",
    "период": "date_raw",
    "месяц": "month_raw",
    "год": "year_raw",
}

# Целевые числовые колонки (возможные). В RAW храним текстом, но пригодится в STAGE.
NUMERIC_TARGETS: Sequence[str] = (
    "quantity", "gross_turnover", "gross_cost", "avg_cost_price", "avg_sell_price",
)

# ------------------------------------------------------------------ #
# Regex / cleaning helpers                                            #
# ------------------------------------------------------------------ #
_pre_ws = re.compile(r"\s+", re.MULTILINE)
_non_word = re.compile(r"[^\w]+", re.UNICODE)
_digits_yyyymm = re.compile(r"(20\d{2})[._-]?([01]?\d)")  # 202409, 2024_09, 2024-9...
_digits_mm_yyyy = re.compile(r"([01]?\d)[._-]?(20\d{2})")  # 09_2024, 9-2024
_year4 = re.compile(r"20\d{2}")


def _clean_header_cell(s: str) -> str:
    """Normalize header cell → snake-ish ascii-ish lower.
    Removes parens, punctuation, condenses underscores.
    """
    if not isinstance(s, str):
        s = "" if s is None else str(s)
    s = s.strip()
    s = _pre_ws.sub(" ", s)
    s = s.replace('"', '').replace("'", "")
    s = re.sub(r"\(.*?\)", "", s)
    s = s.strip()
    s = s.replace(".", "_")
    s = s.replace(" ", "_")
    s = _non_word.sub("_", s)
    s = re.sub(r"_+", "_", s)
    return s.strip("_").lower()


def _rename_using_mapping(df: pd.DataFrame) -> pd.DataFrame:
    cleaned_map = {_clean_header_cell(k): v for k, v in ASHAN_COLUMN_MAPPING.items()}
    new_cols = []
    for c in df.columns:
        ck = _clean_header_cell(c)
        new_cols.append(cleaned_map.get(ck, ck))
    out = df.copy()
    out.columns = new_cols
    return out


def _make_unique_columns(cols: Sequence[str]) -> List[str]:
    seen = {}
    out = []
    for c in cols:
        base = c or "col"
        if base not in seen:
            seen[base] = 1
            out.append(base)
        else:
            seen[base] += 1
            out.append(f"{base}_{seen[base]}")
    return out


def _drop_all_nulls(df: pd.DataFrame) -> pd.DataFrame:
    return df.dropna(axis=0, how="all").dropna(axis=1, how="all")


# ------------------------------------------------------------------ #
# Period / date extraction                                            #
# ------------------------------------------------------------------ #

def _extract_month_year_from_filename(file_name: str) -> Tuple[Optional[int], Optional[int]]:
    """Try to parse month/year from filename tokens.
    Supports RU/EN month names + numeric patterns like 202409, 09_2024.
    Returns (month, year).
    """
    fn = str(file_name).lower()

    # numeric yyyymm?
    m = _digits_yyyymm.search(fn)
    if m:
        year = int(m.group(1))
        month = int(m.group(2))
        if 1 <= month <= 12:
            return month, year

    # numeric mm_yyyy?
    m = _digits_mm_yyyy.search(fn)
    if m:
        month = int(m.group(1))
        year = int(m.group(2))
        if 1 <= month <= 12:
            return month, year

    # token scan
    toks = re.findall(r"[a-zа-я]+|\d{4}", fn)
    year = None
    month = None
    for t in toks:
        if t.isdigit() and len(t) == 4:
            year = int(t)
            continue
        if t in ASHAN_MONTHS:
            month = ASHAN_MONTHS[t]
            continue
        t3 = t[:3]
        if t3 in ASHAN_MONTHS:
            month = ASHAN_MONTHS[t3]
            continue
    return month, year


def _parse_date_str_like(s: str) -> Optional[pd.Timestamp]:
    if not isinstance(s, str):
        s = str(s) if s is not None else ""
    s = s.strip().replace('\xa0', ' ')
    if not s:
        return None
    # Try pandas parser
    try:
        return pd.to_datetime(s, dayfirst=True, errors='raise')
    except Exception:
        pass
    # Try numeric patterns
    m = _digits_yyyymm.search(s)
    if m:
        year = int(m.group(1)); month = int(m.group(2))
        if 1 <= month <= 12:
            return pd.Timestamp(year=year, month=month, day=1)
    m = _digits_mm_yyyy.search(s)
    if m:
        month = int(m.group(1)); year = int(m.group(2))
        if 1 <= month <= 12:
            return pd.Timestamp(year=year, month=month, day=1)
    return None


def _derive_period_columns(df: pd.DataFrame, file_name: str) -> pd.DataFrame:
    """Ensure sale_year, sale_month, sale_date.
    Priority:
      1) Explicit date column in df (date_raw or similar) → parse.
      2) (year_raw, month_raw) columns.
      3) Filename tokens.
    """
    df = df.copy()

    # Step 1: try explicit date columns
    date_col_candidates = [c for c in df.columns if c.lower() in ("date_raw", "date", "дата", "period", "период")]
    dt_obj: Optional[pd.Timestamp] = None
    if date_col_candidates:
        col = date_col_candidates[0]
        sample = df[col].dropna().astype(str).head(1)
        if not sample.empty:
            dt_obj = _parse_date_str_like(sample.iloc[0])
    if dt_obj is not None:
        df['sale_year'] = str(dt_obj.year)
        df['sale_month'] = f"{dt_obj.month:02d}"
        df['sale_date'] = dt_obj.strftime('%Y-%m-%d')
        return df

    # Step 2: month_raw/year_raw columns
    month_raw = None
    year_raw = None
    if 'month_raw' in df.columns:
        sample = df['month_raw'].dropna().astype(str).head(1)
        if not sample.empty:
            # could be text month name or number
            tok = sample.iloc[0].strip().lower()
            if tok.isdigit():
                try:
                    m_int = int(tok)
                    if 1 <= m_int <= 12:
                        month_raw = m_int
                except Exception:
                    pass
            elif tok in ASHAN_MONTHS:
                month_raw = ASHAN_MONTHS[tok]
    if 'year_raw' in df.columns:
        sample = df['year_raw'].dropna().astype(str).head(1)
        if not sample.empty and sample.iloc[0].isdigit():
            year_raw = int(sample.iloc[0])

    if month_raw and year_raw:
        df['sale_year'] = str(year_raw)
        df['sale_month'] = f"{month_raw:02d}"
        df['sale_date'] = f"{year_raw:04d}-{month_raw:02d}-01"
        return df

    # Step 3: filename
    m, y = _extract_month_year_from_filename(file_name)
    if m:
        df['sale_month'] = f"{m:02d}"
    if y:
        df['sale_year'] = str(y)
    if m and y:
        df['sale_date'] = f"{y:04d}-{m:02d}-01"

    return df


# ------------------------------------------------------------------ #
# File readers (CSV + streaming Excel)                               #
# ------------------------------------------------------------------ #

def read_csv_detect_sep(file_path: str, sample_bytes: int = 50_000, default: str = ';') -> str:
    """Detect delimiter by sampling."""
    try:
        with open(file_path, 'r', encoding='utf-8-sig', errors='replace') as f:
            sample = f.read(sample_bytes)
    except Exception:  # pragma: no cover
        return default
    counts = {ch: sample.count(ch) for ch in (';', '\t', ',')}
    sep = max(counts, key=counts.get)
    if counts.get(sep, 0) == 0:
        sep = default
    return sep


def read_csv_stream(file_path: str, chunksize: int) -> Iterator[pd.DataFrame]:
    sep = read_csv_detect_sep(file_path)
    for chunk in pd.read_csv(
        file_path,
        sep=sep,
        dtype='string',
        chunksize=chunksize,
        encoding='utf-8-sig',
        on_bad_lines='warn',
        decimal=',',
        thousands=' ',
        engine='c'
    ):
        yield chunk


def read_excel_stream(file_path: str, sheet_name: str, chunk_size: int) -> Iterator[pd.DataFrame]:
    """Stream rows from an Excel sheet using openpyxl read_only mode."""
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        logger.warning(f"[Ashan] Sheet {sheet_name} not found in {file_path} — skip.")
        return iter(())  # empty iterator
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        return iter(())
    headers = [_clean_header_cell(h) for h in headers]
    chunk: List[Tuple] = []
    for row in rows_iter:
        chunk.append(row)
        if len(chunk) >= chunk_size:
            yield pd.DataFrame(chunk, columns=headers)
            chunk = []
    if chunk:
        yield pd.DataFrame(chunk, columns=headers)


# ------------------------------------------------------------------ #
# Processor                                                           #
# ------------------------------------------------------------------ #
class AshanTableProcessor:
    """Stream-oriented loader of Ashan sales files into RAW schema."""

    MAX_ROWS = 10_000_000  # safety
    CHUNKSIZE = 100_000     # how many rows to read/process at a time
    BATCH_SIZE = 50_000     # how many rows per DB executemany batch
    RAW_SCHEMA = 'raw'

    # Типы колонок override: тк RAW текстовый, но можно уточнить.
    TYPE_OVERRIDES: Dict[str, str] = {
        'sale_year': 'NVARCHAR(4)',
        'sale_month': 'NVARCHAR(2)',
        'sale_date': 'NVARCHAR(10)',  # YYYY-MM-DD
    }

    @classmethod
    def process_file(cls, file_path: str, engine: Engine) -> str:
        start_time = time.time()
        file_name = os.path.basename(file_path)
        base_name = os.path.splitext(file_name)[0]
        table_name = re.sub(r'\W+', '_', base_name.lower())
        logger.info(f"[Ashan] Start processing file {file_name} → raw.{table_name}")

        # Determine input type
        ext = os.path.splitext(file_path)[1].lower()
        sheet_to_iter_map: Dict[str, Iterator[pd.DataFrame]] = {}

        if ext in ('.xlsx', '.xlsm', '.xls'):
            # openpyxl read_only stream per sheet
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            for sh in wb.sheetnames:
                sheet_to_iter_map[sh] = read_excel_stream(file_path, sh, cls.CHUNKSIZE)
        elif ext == '.csv':
            sheet_to_iter_map['data'] = read_csv_stream(file_path, cls.CHUNKSIZE)
        else:
            raise ValueError(f"Неподдерживаемый формат: {file_path}")

        # We'll create table lazily on first non-empty processed chunk.
        table_created = False
        total_rows = 0

        for sh_name, df_iter in sheet_to_iter_map.items():
            logger.info(f"[Ashan] Reading sheet {sh_name}...")
            chunk_idx = 0
            for raw_chunk in df_iter:
                chunk_idx += 1
                if raw_chunk is None or raw_chunk.empty:
                    continue
                logger.debug(f"[Ashan] Sheet {sh_name}: got raw chunk {chunk_idx} with {len(raw_chunk)} rows.")

                # Process chunk
                processed_chunk = cls.process_data_chunk(raw_chunk, file_name, is_first_chunk=(not table_created))
                if processed_chunk.empty:
                    logger.info(f"[Ashan] Processed chunk {chunk_idx} from {sh_name} empty — skip.")
                    continue

                # Create table if needed
                if not table_created:
                    cls.create_table_if_not_exists(processed_chunk, table_name, engine)
                    table_created = True

                # Insert chunk
                inserted = cls.bulk_insert_chunk(processed_chunk, table_name, engine)
                total_rows += inserted
                logger.info(f"[Ashan] Inserted chunk {chunk_idx} from {sh_name}: {inserted} rows (total {total_rows}).")

        if not table_created:
            raise ValueError(f"[Ashan] No data found in file {file_name} — nothing loaded.")

        dur = time.time() - start_time
        logger.info(f"[Ashan] File {file_name} loaded → raw.{table_name}: {total_rows} rows in {dur:.2f}s.")
        return table_name

    # ------------------------------------------------------------------ #
    # Chunk-level processing                                            #
    # ------------------------------------------------------------------ #
    @classmethod
    def process_data_chunk(cls, df: pd.DataFrame, file_name: str, is_first_chunk: bool) -> pd.DataFrame:
        """Normalize & enrich one chunk with ML models."""
        if is_first_chunk:
            df.columns = [_clean_header_cell(c) for c in df.columns]
        df = _rename_using_mapping(df)
        df.columns = _make_unique_columns(df.columns.tolist())
        df = _drop_all_nulls(df)

        # Force string, preserve None for NA
        for col in df.columns:
            try:
                df[col] = df[col].astype('string')
            except Exception:
                df[col] = df[col].astype(str)
        df = df.replace({pd.NA: None, np.nan: None})

        df = _derive_period_columns(df, file_name)

        # Trim values to <=255
        for col in df.columns:
            df[col] = df[col].astype(str).str.slice(0, 255)

        # ---------------------------------------------------------- #
        # ML-модели: enrich продуктами (бренд, вкус и т.д.)          #
        # ---------------------------------------------------------- #
        name_col_candidates = [c for c in df.columns if 'наимен' in c.lower() or 'product' in c.lower()]
        if not name_col_candidates:
            logger.warning(f"[Ashan] Не найдена колонка с названием продукта — enrichment пропущен.")
            return df

        product_col = name_col_candidates[0]
        df['product_name'] = df[product_col]

        # Модели и векторайзеры
        model_dir = "ml_models/product_enrichment"
        model_paths = {
            'flavor': ("flavor_model.pkl", "flavor_vectorizer.pkl"),
            'brand': ("brand_model.pkl", "brand_vectorizer.pkl"),
            'weight': ("weight_model.pkl", "weight_vectorizer.pkl"),
            'product_type': ("type_model.pkl", "type_vectorizer.pkl"),
            # 'package_type': ("package_model.pkl", "package_vectorizer.pkl"),  # опционально
        }

        def load_pickle(path):
            with open(path, "rb") as f:
                return pickle.load(f)

        for col_name, (model_file, vec_file) in model_paths.items():
            try:
                model = load_pickle(os.path.join(model_dir, model_file))
                vectorizer = load_pickle(os.path.join(model_dir, vec_file))
                vec = vectorizer.transform(df['product_name'].astype(str))
                df[col_name] = model.predict(vec)
            except Exception as e:
                logger.warning(f"[Ashan] Не удалось обогатить '{col_name}': {e}")

        return df


    # ------------------------------------------------------------------ #
    # Table DDL helpers                                                 #
    # ------------------------------------------------------------------ #
    @classmethod
    def create_table_if_not_exists(cls, df: pd.DataFrame, table_name: str, engine: Engine):
        schema = cls.RAW_SCHEMA
        # Clean column names for DDL
        cols = [c.replace(']', '').replace('[', '') for c in df.columns]
        cols_ddl = []
        for c in cols:
            sql_type = cls.TYPE_OVERRIDES.get(c, 'NVARCHAR(255)')
            cols_ddl.append(f"[{c}] {sql_type}")
        create_sql = f"""
        IF NOT EXISTS (
            SELECT 1 FROM INFORMATION_SCHEMA.TABLES
            WHERE TABLE_SCHEMA = '{schema}' AND TABLE_NAME = '{table_name}'
        )
        BEGIN
            CREATE TABLE [{schema}].[{table_name}] (
                {', '.join(cols_ddl)}
            )
        END
        """
        with engine.begin() as conn:
            logger.info(f"[Ashan] Ensure table {schema}.{table_name} exists.")
            conn.execute(text(create_sql))
            logger.info(f"[Ashan] Table {schema}.{table_name} ready.")

    @classmethod
    def _fetch_table_columns(cls, table_name: str, engine: Engine) -> List[str]:
        schema = cls.RAW_SCHEMA
        with engine.connect() as conn:
            result = conn.execute(text("""
                SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_SCHEMA=:sch AND TABLE_NAME=:tbl
                ORDER BY ORDINAL_POSITION
            """), {"sch": schema, "tbl": table_name})
            return [r[0] for r in result]

    @classmethod
    def _add_columns(cls, new_columns: Sequence[str], table_name: str, engine: Engine):
        if not new_columns:
            return
        schema = cls.RAW_SCHEMA
        with engine.begin() as conn:
            for c in new_columns:
                sql_type = cls.TYPE_OVERRIDES.get(c, 'NVARCHAR(255)')
                logger.info(f"[Ashan] ALTER TABLE add column {c} {sql_type}.")
                conn.execute(text(f"ALTER TABLE [{schema}].[{table_name}] ADD [{c}] {sql_type}"))

    # ------------------------------------------------------------------ #
    # Insert                                                            #
    # ------------------------------------------------------------------ #
    @classmethod
    def bulk_insert_chunk(cls, df: pd.DataFrame, table_name: str, engine: Engine) -> int:
        """Insert one processed chunk into raw.table.
        Returns number of inserted rows.
        """
        if df.empty:
            return 0

        schema = cls.RAW_SCHEMA
        # Ensure DB has all required columns (case-insensitive compare?)
        table_columns = cls._fetch_table_columns(table_name, engine)
        cols_lower = {c.lower(): c for c in table_columns}
        missing = []
        for c in df.columns:
            if c.lower() not in cols_lower:
                missing.append(c)
        if missing:
            cls._add_columns(missing, table_name, engine)
            table_columns = cls._fetch_table_columns(table_name, engine)

        # Align df
        df = df.reindex(columns=table_columns, fill_value=None)

        # Prepare data tuples
        # Avoid per-cell Python conversions in loop: vectorize to numpy object array
        arr = df.to_numpy(dtype=object, copy=False)
        # Replace nan-likes with None; str slice already done
        mask = pd.isna(arr)
        if mask.any():
            arr[mask] = None
        # Convert to tuples
        data = list(map(tuple, arr.tolist()))
        del arr

        # Insert batched
        cols_sql = ', '.join(f'[{c}]' for c in table_columns)
        params_sql = ', '.join(['?'] * len(table_columns))
        insert_sql = f"INSERT INTO [{schema}].[{table_name}] ({cols_sql}) VALUES ({params_sql})"

        inserted_total = 0
        raw_conn = engine.raw_connection()
        try:
            cursor = raw_conn.cursor()
            cursor.fast_executemany = True
            # Optionally: cursor.setinputsizes(...)

            bs = cls.BATCH_SIZE
            for i in range(0, len(data), bs):
                batch = data[i:i+bs]
                try:
                    cursor.executemany(insert_sql, batch)
                    inserted_total += len(batch)
                except Exception as e:  # pragma: no cover
                    logger.error(f"[Ashan] batch insert error rows {i}-{i+len(batch)}: {e}", exc_info=True)
                    raise
            raw_conn.commit()
        finally:
            raw_conn.close()
        return inserted_total


# ------------------------------------------------------------------ #
# Public wrapper                                                      #
# ------------------------------------------------------------------ #
def create_ashan_table_and_upload(file_path: str, engine: Engine) -> str:
    try:
        return AshanTableProcessor.process_file(file_path, engine)
    except Exception as e:  # pragma: no cover
        logger.error(f"[Ashan] Critical error processing {file_path}: {e}", exc_info=True)
        raise


# ------------------------------------------------------------------ #
# CLI usage (optional)                                                #
# ------------------------------------------------------------------ #
if __name__ == '__main__':  # pragma: no cover
    import argparse
    logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
    parser = argparse.ArgumentParser(description='Load Ashan file into RAW schema.')
    parser.add_argument('path', help='Path to input file (.xlsx/.csv).')
    parser.add_argument('--conn', help='SQLAlchemy connection string mssql+pyodbc://...', required=True)
    parser.add_argument('--schema', default=AshanTableProcessor.RAW_SCHEMA, help='Target schema (default raw).')
    parser.add_argument('--chunksize', type=int, default=AshanTableProcessor.CHUNKSIZE)
    parser.add_argument('--batchsize', type=int, default=AshanTableProcessor.BATCH_SIZE)
    args = parser.parse_args()

    from sqlalchemy import create_engine
    AshanTableProcessor.RAW_SCHEMA = args.schema
    AshanTableProcessor.CHUNKSIZE = args.chunksize
    AshanTableProcessor.BATCH_SIZE = args.batchsize

    eng = Engine = create_engine(args.conn, fast_executemany=True)
    create_ashan_table_and_upload(args.path, eng)
